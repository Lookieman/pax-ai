"""
Golden Ground Truth Converter
=============================
Converts manually labelled Excel ground truth files (Salmonella and E. coli)
into standardised JSON format for GEPA evaluation.

Paths are loaded from config.py (GOLDEN_GT_INPUT_PATH and GOLDEN_GT_PATH).

Author: Luqman (AI6129 Pathogen Tracking Project)
Date:   February 2026

Dependencies: openpyxl (pip install openpyxl)
"""

import os
import re
import json
import logging
from pathlib import Path

import openpyxl

from config import cfg  # #changed - centralised configuration

# ===========================================================================
# Configuration (from config.py)
# ===========================================================================
INPUT_DIR = cfg.GOLDEN_GT_INPUT_PATH   # #changed
OUTPUT_DIR = cfg.GOLDEN_GT_PATH        # #changed

# Salmonella file name (single file, all PMCIDs)
SALMONELLA_FILE = cfg.SALMONELLA_FILE          # #changed
ECOLI_ISOLATES_SUFFIX = cfg.ECOLI_ISOLATES_SUFFIX  # #changed
ECOLI_OTHERS_SUFFIX = cfg.ECOLI_OTHERS_SUFFIX      # #changed

# Logging setup
logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)s | %(message)s"
)
logger = logging.getLogger(__name__)

# ===========================================================================
# Interpretation mappings
# ===========================================================================
INTERP_WORD_MAP = {
    "resistant": "R",
    "susceptible": "S",
    "intermediate": "I",
    "sensible": "S",       # French/Spanish for susceptible
}

VALID_INTERP = {"S", "I", "R"}


# ===========================================================================
# Legend loader
# ===========================================================================
def load_legend(filepath):
    """
    Load the antibiotic abbreviation legend from the 'Legend' sheet
    of the Salmonella Excel file.

    Expected columns: Abbreviation, Full_Name
    Returns a dict: abbreviation (uppercase) -> full_name (lowercase)
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    if "Legend" not in wb.sheetnames:
        logger.error("No 'Legend' sheet found in %s", filepath)
        raise ValueError("Missing 'Legend' sheet in Salmonella file")

    ws = wb["Legend"]
    legend = {}
    row_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        abbrev = row[0]
        full_name = row[1]
        if abbrev is None or full_name is None:
            continue
        abbrev_clean = str(abbrev).strip().upper()
        full_name_clean = str(full_name).strip().lower()
        legend[abbrev_clean] = full_name_clean
        row_count += 1

    wb.close()
    logger.info("Loaded %d legend entries", row_count)
    return legend


# ===========================================================================
# Superscript normalisation
# ===========================================================================
def normalise_superscript(text):
    """
    Normalise superscript variants in AST keywords.

    NR^{CIP}, NR^CIP, NR^{-CIP}, NR-CIP -> NRCIP
    NAL^{CIP}, NAL^CIP, NAL^{-CIP}, NAL-CIP -> NALCIP
    SXT^{CIP}, SXT^CIP, SXT-CIP -> SXTCIP
    """
    # Remove ^, {, }, and any - immediately after ^
    cleaned = text.strip()
    cleaned = cleaned.replace("^{-", "").replace("^{", "").replace("^", "")
    cleaned = cleaned.replace("{", "").replace("}", "")

    # Handle remaining dash between base and CIP for known patterns
    for base in ["NR", "NAL", "SXT"]:
        pattern = base + "-CIP"
        if cleaned.upper() == pattern:
            cleaned = base + "CIP"

    return cleaned.upper()


# ===========================================================================
# Legend lookup
# ===========================================================================
def lookup_legend(abbreviation, legend, pmcid=""):
    """
    Look up an abbreviation in the legend with fallback logic:
    1. Exact match (uppercase)
    2. Strip disc amount (e.g. AMC-30 -> AMC)
    3. Strip decimal disc amount (e.g. SSS-.25 -> SSS)
    4. Case-insensitive match

    Returns the full antibiotic name, or the original abbreviation if not found.
    Logs a warning for unresolved abbreviations.
    """
    abbrev_upper = abbreviation.strip().upper()

    # Step 1: exact match
    if abbrev_upper in legend:
        return legend[abbrev_upper]

    # Step 2: strip numeric suffix (e.g. AMC-30 -> AMC)
    stripped = re.sub(r"-\.?\d+\.?\d*$", "", abbrev_upper)
    if stripped != abbrev_upper and stripped in legend:
        return legend[stripped]

    # Step 3: case-insensitive scan
    for key, value in legend.items():
        if key == abbrev_upper:
            return value

    # Not found
    logger.warning(
        "[%s] Unresolved abbreviation: '%s' (stripped: '%s')",
        pmcid, abbreviation, stripped
    )
    return abbreviation.strip()


# ===========================================================================
# Special keyword handling
# ===========================================================================
SPECIAL_KEYWORDS = {
    "NR", "NRCIP", "NALCIP", "SXTCIP", "PANSUSCEPTIBLE", "SENSIBLE"
}


def handle_special_keyword(raw_value, serotype, legend, pmcid=""):
    """
    Handle special AST keywords (NR, NRCIP, NALCIP, etc.).
    Returns a list of AST_Data dicts, or None if not a special keyword.
    """
    normalised = normalise_superscript(raw_value)

    if normalised not in SPECIAL_KEYWORDS:
        return None

    if normalised in ("NR", "PANSUSCEPTIBLE", "SENSIBLE"):
        # Pan-susceptible: no specific antibiotics to list
        return [{
            "Serotype": serotype,
            "Antibiotics": [{
                "Name": "Pansusceptible",
                "MIC": "null",
                "Interpretation": "S"
            }]
        }]

    if normalised == "NRCIP":
        return [{
            "Serotype": serotype,
            "Antibiotics": [{
                "Name": "ciprofloxacin",
                "MIC": "null",
                "Interpretation": "I"
            }]
        }]

    if normalised == "NALCIP":
        return [{
            "Serotype": serotype,
            "Antibiotics": [
                {"Name": "nalidixic acid", "MIC": "null", "Interpretation": "R"},
                {"Name": "ciprofloxacin", "MIC": "null", "Interpretation": "I"},
            ]
        }]

    if normalised == "SXTCIP":
        sxt_name = legend.get("SXT", "sulfamethoxazole/trimethoprim")
        return [{
            "Serotype": serotype,
            "Antibiotics": [
                {"Name": sxt_name, "MIC": "null", "Interpretation": "R"},
                {"Name": "ciprofloxacin", "MIC": "null", "Interpretation": "I"},
            ]
        }]

    return None


# ===========================================================================
# Format detection
# ===========================================================================
def detect_format(ast_string):
    """
    Detect the AST_Data format using the decision tree from AST_Converter.md.
    Returns a format number (1-9) or 0 for special keywords, -1 for unknown.
    """
    text = ast_string.strip()

    # Check for JSON (E.coli format)
    if text.startswith("{") or text.startswith("["):
        return 100  # JSON format

    # Check for special keywords
    normalised = normalise_superscript(text)
    if normalised in SPECIAL_KEYWORDS:
        return 0

    # Check for free text / manual review entries
    if any(phrase in text.lower() for phrase in [
        "pls refer", "please refer", "the mlst", "refer to"
    ]):
        return -1

    # Format 5: "ABBREV: FullName - Interpretation" (newline-separated)
    if "\n" in text and ": " in text and " - " in text:
        lines = [l.strip() for l in text.split("\n") if l.strip()]
        # Check if lines match ABBREV: Name - Interp pattern
        if all(re.match(r"^[A-Z]{1,6}:\s+.+\s+-\s+\w+", l) for l in lines[:3]):
            return 5

    # Format 6: "FullName - Susceptible/Resistant/Intermediate" (newline-separated)
    if "\n" in text:
        lines = [l.strip() for l in text.split("\n") if l.strip()]
        if any(
            l.endswith("Susceptible") or l.endswith("Resistant") or l.endswith("Intermediate")
            for l in lines[:3]
        ):
            return 6

    # Format 3: "FullName: S/I/R" (colon-space before single letter interp)
    if ": " in text:
        entries = [e.strip() for e in text.split(",") if e.strip()]
        # Check if most entries match "Something: S/I/R"
        colon_matches = sum(
            1 for e in entries
            if re.search(r":\s*[SIR]$", e.strip())
        )
        if colon_matches > len(entries) * 0.5:
            # Distinguish Format 3 from Format 5
            # Format 5 has abbreviation before colon (1-4 uppercase letters)
            first_before_colon = entries[0].split(":")[0].strip()
            if re.match(r"^[A-Z]{1,6}$", first_before_colon):
                # Could be Format 5 single-line variant, but Format 5 is typically
                # multi-line. Treat comma-separated colon format as Format 3.
                pass
            return 3

    # Format 6 single-line variant: "Name - Resistant, Name - Susceptible"
    if " - Susceptible" in text or " - Resistant" in text or " - Intermediate" in text:
        return 6

    # Format 1: entries with disc amounts like AMC-30-R
    entries = [e.strip() for e in text.split(",") if e.strip()]
    disc_pattern = re.compile(r"^[A-Z]{1,6}-\.?\d+\.?\d*-[SIR]$", re.IGNORECASE)
    if any(disc_pattern.match(e) for e in entries):
        return 1

    # Format 2: entries like AMC-S (short abbreviation, single hyphen, S/I/R)
    compact_pattern = re.compile(r"^[A-Z]{1,8}-[SIR]$", re.IGNORECASE)
    if all(compact_pattern.match(e) for e in entries if e):
        return 2

    # Format 8: "Abbrev-R," with mixed case and trailing commas
    format8_pattern = re.compile(r"^[A-Za-z]{2,6}-R$")
    clean_entries = [e.strip().rstrip(",") for e in text.split(",") if e.strip()]
    # Format 8 also has space-delimited multi-drug tokens (e.g. "Gen Kan-R")
    if clean_entries and all(
        format8_pattern.match(token.strip().split()[-1])
        for token in clean_entries if token.strip()
    ):
        return 8

    # Format 10: Full-name hyphen-separated list, all resistant  #changed
    # E.g. "sulfisoxazole-ampicillin-ticarcillin/clavulanic acid constant 2"  #changed
    # Detected by: hyphens present, no commas, contains "constant" OR  #changed
    # any segment is a long full drug name (>6 chars)  #changed
    if "-" in text and "," not in text:  #changed
        cleaned = re.sub(r"\s+constant\s+\d+$", "", text.strip())  #changed
        if "constant" in text.lower() or "/" in cleaned:  #changed
            return 10  #changed

    # Format 9: Hyphen-separated mixed-case abbreviations (Amp-Cip-Sul-Tet, FOX-TET-AUG2)
    if "-" in text and "," not in text and " " not in text.strip():  #changed
        parts = text.split("-")
        # All parts should be short abbreviations (2-6 chars, allowing trailing digits)  #changed
        if all(re.match(r"^[A-Za-z]{2,5}\d?$", p.strip()) for p in parts if p.strip()):  #changed
            return 9

    # Format 4: "FullName-S" with full drug names
    format4_pattern = re.compile(r"^.+-[SIR]$")
    if entries and all(format4_pattern.match(e) for e in entries):
        return 4

    # Format 7: abbreviation-only lists (all resistant, no S/I/R)
    # Could be comma-separated or space-separated
    # Also handles mixed full-name + abbreviation entries (e.g. "Sulfonamides, AMP")  #changed
    if "," in text:
        tokens = [t.strip() for t in text.split(",") if t.strip()]
    else:
        tokens = text.split()

    if tokens and all(  #changed
        re.match(r"^[A-Za-z]{1,8}\d?$", t, re.IGNORECASE)  # short abbreviation  #changed
        or (re.match(r"^[A-Za-z]+$", t) and len(t) > 3)     # full drug name  #changed
        for t in tokens  #changed
    ):  #changed
        return 7

    return -1


# ===========================================================================
# Format parsers
# ===========================================================================
def parse_format_1(ast_string, serotype, legend, pmcid=""):
    """Format 1: Compact abbreviation with disc amounts. E.g. AMC-30-R, AM-10-S"""
    entries = [e.strip() for e in ast_string.split(",") if e.strip()]
    antibiotics = []

    for entry in entries:
        # Interpretation is the last character after the final hyphen
        parts = entry.rsplit("-", 1)
        if len(parts) != 2 or parts[1] not in VALID_INTERP:
            logger.warning("[%s] Format 1: cannot parse entry '%s'", pmcid, entry)
            continue

        interp = parts[1]
        abbrev_with_disc = parts[0]  # e.g. AMC-30, SSS-.25, or just AMC

        full_name = lookup_legend(abbrev_with_disc, legend, pmcid)
        antibiotics.append({
            "Name": full_name,
            "MIC": "null",
            "Interpretation": interp
        })

    if not antibiotics:
        return []

    return [{"Serotype": serotype, "Antibiotics": antibiotics}]


def parse_format_2(ast_string, serotype, legend, pmcid=""):
    """Format 2: Compact abbreviation without disc amounts. E.g. AMC-S, AMP-R"""
    entries = [e.strip() for e in ast_string.split(",") if e.strip()]
    antibiotics = []

    for entry in entries:
        parts = entry.rsplit("-", 1)
        if len(parts) != 2 or parts[1] not in VALID_INTERP:
            logger.warning("[%s] Format 2: cannot parse entry '%s'", pmcid, entry)
            continue

        interp = parts[1]
        abbrev = parts[0].strip()
        full_name = lookup_legend(abbrev, legend, pmcid)
        antibiotics.append({
            "Name": full_name,
            "MIC": "null",
            "Interpretation": interp
        })

    if not antibiotics:
        return []

    return [{"Serotype": serotype, "Antibiotics": antibiotics}]


def parse_format_3(ast_string, serotype, legend, pmcid=""):
    """Format 3: FullName: Interpretation. E.g. Ampicillin: R, Cefoxitin: S"""
    entries = [e.strip() for e in ast_string.split(",") if e.strip()]
    antibiotics = []

    for entry in entries:
        # Split on the last colon
        parts = entry.rsplit(":", 1)
        if len(parts) != 2:
            logger.warning("[%s] Format 3: cannot parse entry '%s'", pmcid, entry)
            continue

        name = parts[0].strip()
        interp_raw = parts[1].strip().upper()

        if interp_raw not in VALID_INTERP:
            # Check for word-form interpretations
            interp_raw = INTERP_WORD_MAP.get(interp_raw.lower(), interp_raw)

        if interp_raw not in VALID_INTERP:
            logger.warning(
                "[%s] Format 3: invalid interpretation '%s' for '%s'",
                pmcid, parts[1].strip(), name
            )
            continue

        antibiotics.append({
            "Name": name.title(),
            "MIC": "null",
            "Interpretation": interp_raw
        })

    if not antibiotics:
        return []

    return [{"Serotype": serotype, "Antibiotics": antibiotics}]


def parse_format_4(ast_string, serotype, legend, pmcid=""):
    """Format 4: FullName-Interpretation (hyphen delimiter). E.g. Florfenicol-S"""
    # Split on comma with optional multiple spaces
    entries = [e.strip() for e in re.split(r",\s+", ast_string) if e.strip()]
    antibiotics = []

    for entry in entries:
        # Parse from the right: last hyphen before a single S/I/R
        match = re.match(r"^(.+)-([SIR])$", entry)
        if not match:
            logger.warning("[%s] Format 4: cannot parse entry '%s'", pmcid, entry)
            continue

        name = match.group(1).strip()
        interp = match.group(2)
        antibiotics.append({
            "Name": name.title(),
            "MIC": "null",
            "Interpretation": interp
        })

    if not antibiotics:
        return []

    return [{"Serotype": serotype, "Antibiotics": antibiotics}]


def parse_format_5(ast_string, serotype, legend, pmcid=""):
    """Format 5: ABBREV: FullName - Interpretation (one per line)."""
    lines = [l.strip() for l in ast_string.split("\n") if l.strip()]
    antibiotics = []

    for line in lines:
        # Pattern: ABBREV: FullName - Resistant/Susceptible/Intermediate
        match = re.match(r"^[A-Z]{1,6}:\s*(.+?)\s*-\s*(\w+)$", line)
        if not match:
            logger.warning("[%s] Format 5: cannot parse line '%s'", pmcid, line)
            continue

        name = match.group(1).strip()
        interp_word = match.group(2).strip().lower()
        interp = INTERP_WORD_MAP.get(interp_word, interp_word.upper())

        if interp not in VALID_INTERP:
            logger.warning(
                "[%s] Format 5: invalid interpretation '%s' for '%s'",
                pmcid, interp_word, name
            )
            continue

        antibiotics.append({
            "Name": name.title(),
            "MIC": "null",
            "Interpretation": interp
        })

    if not antibiotics:
        return []

    return [{"Serotype": serotype, "Antibiotics": antibiotics}]


def parse_format_6(ast_string, serotype, legend, pmcid=""):
    """Format 6: FullName - Susceptible/Resistant/Intermediate."""
    # Could be newline-separated or comma-separated
    if "\n" in ast_string:
        entries = [l.strip() for l in ast_string.split("\n") if l.strip()]
    else:
        # Comma-separated but with " - Resistant" etc.
        entries = [e.strip() for e in ast_string.split(",") if e.strip()]

    antibiotics = []

    for entry in entries:
        # Split from the right on " - "
        parts = entry.rsplit(" - ", 1)
        if len(parts) != 2:
            logger.warning("[%s] Format 6: cannot parse entry '%s'", pmcid, entry)
            continue

        name = parts[0].strip()
        interp_word = parts[1].strip().lower()

        # Handle "constant 2" suffix
        name = re.sub(r"\s+constant\s+\d+$", "", name)

        interp = INTERP_WORD_MAP.get(interp_word, interp_word.upper())
        if interp not in VALID_INTERP:
            logger.warning(
                "[%s] Format 6: invalid interpretation '%s' for '%s'",
                pmcid, interp_word, name
            )
            continue

        antibiotics.append({
            "Name": name.title(),
            "MIC": "null",
            "Interpretation": interp
        })

    if not antibiotics:
        return []

    return [{"Serotype": serotype, "Antibiotics": antibiotics}]


def parse_format_7(ast_string, serotype, legend, pmcid=""):
    """Format 7: Abbreviation-only list (all resistant). E.g. FIS, KAN, STR"""
    text = ast_string.strip()

    # Detect delimiter: comma or space
    if "," in text:
        tokens = [t.strip() for t in text.split(",") if t.strip()]
    else:
        tokens = [t.strip() for t in text.split() if t.strip()]

    antibiotics = []
    for token in tokens:
        # Check if it is a full drug name (contains lowercase letters and length > 5)
        if re.match(r"^[A-Za-z]+$", token) and len(token) > 5 and token[0].isupper():
            # Full drug name like "Sulfonamides", "Tetracycline"
            name = token.title()
        else:
            name = lookup_legend(token, legend, pmcid)

        antibiotics.append({
            "Name": name,
            "MIC": "null",
            "Interpretation": "R"
        })

    if not antibiotics:
        return []

    return [{"Serotype": serotype, "Antibiotics": antibiotics}]


def parse_format_8(ast_string, serotype, legend, pmcid=""):
    """
    Format 8: Abbreviation-R (resistant list with trailing commas).
    E.g. Amp-R, Chl-R, Gen Kan-R, Nal Fis-R,
    """
    # Split on comma, strip whitespace and trailing commas
    raw_tokens = [t.strip().rstrip(",") for t in ast_string.split(",") if t.strip()]
    antibiotics = []

    for token in raw_tokens:
        if not token:
            continue

        # Check for space-delimited multi-drug tokens: "Gen Kan-R"
        if " " in token:
            # Split on space; last part has the -R suffix
            sub_parts = token.split()
            for part in sub_parts:
                # Strip -R suffix if present
                abbrev = re.sub(r"-R$", "", part, flags=re.IGNORECASE).strip()
                if abbrev:
                    full_name = lookup_legend(abbrev, legend, pmcid)
                    antibiotics.append({
                        "Name": full_name,
                        "MIC": "null",
                        "Interpretation": "R"
                    })
        else:
            # Standard Abbrev-R token
            abbrev = re.sub(r"-R$", "", token, flags=re.IGNORECASE).strip()
            if abbrev:
                full_name = lookup_legend(abbrev, legend, pmcid)
                antibiotics.append({
                    "Name": full_name,
                    "MIC": "null",
                    "Interpretation": "R"
                })

    if not antibiotics:
        return []

    return [{"Serotype": serotype, "Antibiotics": antibiotics}]


def parse_format_9(ast_string, serotype, legend, pmcid=""):
    """Format 9: Mixed abbreviation (all resistant). E.g. Amp-Cip-Sul-Tet"""
    parts = [p.strip() for p in ast_string.split("-") if p.strip()]
    antibiotics = []

    for part in parts:
        full_name = lookup_legend(part, legend, pmcid)
        antibiotics.append({
            "Name": full_name,
            "MIC": "null",
            "Interpretation": "R"
        })

    if not antibiotics:
        return []

    return [{"Serotype": serotype, "Antibiotics": antibiotics}]


def parse_format_10(ast_string, serotype, legend, pmcid=""):  #changed
    """  #changed
    Format 10: Full-name hyphen-separated list, all resistant.  #changed
    E.g. 'sulfisoxazole-ampicillin-ticarcillin/clavulanic acid constant 2'  #changed
    """  #changed
    text = ast_string.strip()  #changed
    # Strip "constant N" suffix  #changed
    text = re.sub(r"\s+constant\s+\d+$", "", text)  #changed

    # Split on hyphens, but preserve compound names with "/"  #changed
    # Strategy: split on "-", then re-join fragments that contain "/" or  #changed
    # are clearly part of a compound name (e.g. "clavulanic acid")  #changed
    raw_parts = text.split("-")  #changed
    drugs = []  #changed
    i = 0  #changed
    while i < len(raw_parts):  #changed
        part = raw_parts[i].strip()  #changed
        # If this part or the next contains "/" or spaces that suggest  #changed
        # it is a continuation (e.g. "ticarcillin/" + "clavulanic acid")  #changed
        # Check if part ends with "/" -> merge with next  #changed
        if part.endswith("/") and i + 1 < len(raw_parts):  #changed
            part = part + raw_parts[i + 1].strip()  #changed
            i += 1  #changed
        # Check if part contains "/" -> it is a compound drug already  #changed
        # Check if part contains a space and looks like a drug name  #changed
        # (not a short abbreviation)  #changed
        if part:  #changed
            drugs.append(part)  #changed
        i += 1  #changed

    # Post-process: re-join fragments where a "/" split across hyphens  #changed
    # E.g. ["ticarcillin/clavulanic acid"] is already correct  #changed
    # But ["ticarcillin", "clavulanic acid"] needs joining  #changed
    # Heuristic: if a fragment starts lowercase and the previous  #changed
    # fragment does not end with a letter, join them  #changed
    merged = []  #changed
    for drug in drugs:  #changed
        if (merged  #changed
                and not re.match(r"^[A-Z]", drug)  #changed
                and " " in drug  #changed
                and "/" in merged[-1]):  #changed
            merged[-1] = merged[-1] + drug  #changed
        else:  #changed
            merged.append(drug)  #changed

    antibiotics = []  #changed
    for drug in merged:  #changed
        antibiotics.append({  #changed
            "Name": drug.strip().title(),  #changed
            "MIC": "null",  #changed
            "Interpretation": "R"  #changed
        })  #changed

    if not antibiotics:  #changed
        return []  #changed

    return [{"Serotype": serotype, "Antibiotics": antibiotics}]  #changed
# ===========================================================================
def parse_ast_data(ast_string, serotype, legend, pmcid=""):
    """
    Master AST parser. Detects format and dispatches to the appropriate parser.
    Returns a list of AST_Data dicts matching the target JSON schema.
    """
    if ast_string is None or str(ast_string).strip() == "":
        return []

    text = str(ast_string).strip()

    # Check for JSON (E.coli pre-formed JSON)
    if text.startswith("{") or text.startswith("["):
        try:
            parsed = json.loads(text)
            if isinstance(parsed, dict):
                return [parsed]
            elif isinstance(parsed, list):
                return parsed
        except json.JSONDecodeError:
            logger.warning("[%s] Failed to parse JSON AST_Data: %s...", pmcid, text[:80])
            return []

    # Check for special keywords first
    special_result = handle_special_keyword(text, serotype, legend, pmcid)
    if special_result is not None:
        return special_result

    # Detect format
    fmt = detect_format(text)

    # Dispatch to parser
    parser_map = {
        1: parse_format_1,
        2: parse_format_2,
        3: parse_format_3,
        4: parse_format_4,
        5: parse_format_5,
        6: parse_format_6,
        7: parse_format_7,
        8: parse_format_8,
        9: parse_format_9,
        10: parse_format_10,  #changed
    }

    if fmt == -1:
        logger.warning(
            "[%s] FLAGGED FOR MANUAL REVIEW - unrecognised AST format: '%s'",
            pmcid, text[:100]
        )
        return []

    parser_func = parser_map.get(fmt)
    if parser_func is None:
        logger.warning("[%s] No parser for detected format %d: '%s'", pmcid, fmt, text[:100])
        return []

    logger.info("[%s] Detected Format %d for: '%s...'", pmcid, fmt, text[:60])
    return parser_func(text, serotype, legend, pmcid)


# ===========================================================================
# Utility functions
# ===========================================================================
def split_to_list(value):
    """Split a comma-separated string into a list, or return empty list if null."""
    if value is None or str(value).strip() == "":
        return []
    return [v.strip() for v in str(value).split(",") if v.strip()]


def extract_years(value):
    """
    Extract year(s) from a date string.
    Handles: 2015, February 2000, 2009 Oct, etc.
    Returns a list of integers.
    """
    if value is None or str(value).strip() == "":
        return []

    text = str(value).strip()

    # Find all 4-digit numbers that look like years (1900-2099)
    years = re.findall(r"\b(19\d{2}|20\d{2})\b", text)
    return [int(y) for y in years]


def safe_string(value):
    """Convert a value to a trimmed string, or return empty string if null."""
    if value is None:
        return ""
    text = str(value).strip()
    text = normalise_dashes(text)  #changed
    return text


def normalise_dashes(text):  #changed
    """  #changed
    Normalise Unicode dash variants to standard hyphen (U+002D).  #changed
    Fixes matching issues where en-dash (U+2013) or em-dash (U+2014)  #changed
    in source data prevents string comparison with regular hyphens.  #changed
    E.g. 'XJ10\u201314' (en-dash) -> 'XJ10-14' (regular hyphen)  #changed
    """  #changed
    # En-dash (U+2013), Em-dash (U+2014), Figure dash (U+2012),  #changed
    # Horizontal bar (U+2015), Minus sign (U+2212)  #changed
    dash_chars = "\u2012\u2013\u2014\u2015\u2212"  #changed
    for ch in dash_chars:  #changed
        text = text.replace(ch, "-")  #changed
    return text  #changed


# ===========================================================================
# Salmonella converter
# ===========================================================================
def convert_salmonella(filepath, legend):
    """
    Convert the single Salmonella Excel file into a dict of PMCID -> JSON structure.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    # Use first sheet (excludes Legend sheet) #changed
    data_sheets = [s for s in wb.sheetnames if s.lower() != "legend"]  #changed
    ws = wb[data_sheets[0]]  #changed

    # Build header index map
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {}
    for i, h in enumerate(headers):
        if h is not None:
            col_idx[h.strip()] = i

    # Validate required columns exist
    required = ["PMCID", "isolate_id", "AST_Data"]
    for col in required:
        if col not in col_idx:
            raise ValueError(f"Missing required column '{col}' in Salmonella file. Found: {list(col_idx.keys())}")

    # Process rows
    results = {}
    current_pmcid = None
    row_count = 0
    flagged_rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_count += 1

        # Get PMCID (should be filled down on every row)
        pmcid_val = row[col_idx["PMCID"]] if col_idx["PMCID"] < len(row) else None
        if pmcid_val is not None and str(pmcid_val).strip():
            current_pmcid = str(pmcid_val).strip()

        if current_pmcid is None:
            logger.warning("Row %d has no PMCID, skipping", row_count + 1)
            continue

        # Initialise PMCID structure if new
        if current_pmcid not in results:
            results[current_pmcid] = {
                "pmcid": current_pmcid,
                "isolate_without_linking": [],
                "no_isolates_only_assayinformation": {},
                "isolates_with_linking": []
            }

        # Helper to get column value safely
        def get_col(name):
            idx = col_idx.get(name)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        # Determine category (default: linked)
        category = safe_string(get_col("category")).lower()
        if category == "":
            category = "linked"

        # Extract fields
        isolate_id = safe_string(get_col("isolate_id"))
        isolate_source = safe_string(get_col("isolate_source"))
        isolate_date = extract_years(get_col("isolate_date"))
        isolate_country = safe_string(get_col("isolate_country"))
        serotype_raw = safe_string(get_col("Serotype"))
        serotype_list = split_to_list(serotype_raw) if serotype_raw else []
        mlst_list = split_to_list(get_col("MLST"))
        ast_raw = get_col("AST_Data")
        spi_list = split_to_list(get_col("SPI"))
        amr_list = split_to_list(get_col("AMR"))
        plasmid_list = split_to_list(get_col("Plasmid"))
        snp_list = split_to_list(get_col("SNP"))
        virulence_list = split_to_list(get_col("Virulence Genes"))
        accession_list = split_to_list(get_col("Accession Numbers"))
        accession_merge = split_to_list(get_col("Accession Numbers_merge"))

        # Parse AST_Data
        serotype_for_ast = serotype_list[0] if serotype_list else None
        ast_data = parse_ast_data(ast_raw, serotype_for_ast, legend, current_pmcid)

        # Build the isolate record
        isolate = {
            "isolate_id": isolate_id,
            "isolate_source": isolate_source,
            "isolate_date": isolate_date,
            "isolate_country": isolate_country,
            "serotype": serotype_list,
            "mlst": mlst_list,
            "ast_data": ast_data,
            "spi": spi_list,
            "amr": amr_list,
            "plasmid": plasmid_list,
            "snp": snp_list,
            "virulence_genes": virulence_list,
        }

        # Collect all accession numbers at PMCID level into merge_accession_number  #changed
        all_accessions = accession_list + accession_merge  #changed
        if all_accessions:  #changed
            results[current_pmcid].setdefault("_merge_accession_number", []).extend(all_accessions)  #changed

        # Route to appropriate section based on category
        # Accept both short forms (linked/unlinked/assay_only) and  #changed
        # full forms (isolate_without_linking/no_isolates_only_assayinformation)  #changed
        if category in ("linked", "isolates_with_linking"):  #changed
            results[current_pmcid]["isolates_with_linking"].append(isolate)
        elif category in ("unlinked", "isolate_without_linking"):  #changed
            results[current_pmcid]["isolate_without_linking"].append(isolate)
        elif category in ("assay_only", "no_isolates_only_assayinformation"):  #changed
            # For assay_only, store as a flat dict (aggregate data, no isolate ID)
            # Merge AST and other assay data into the dict
            existing = results[current_pmcid]["no_isolates_only_assayinformation"]
            if "ast_data" not in existing:
                existing["ast_data"] = []
            existing["ast_data"].extend(ast_data)
            if amr_list:
                existing.setdefault("amr", []).extend(amr_list)
            if serotype_list:
                existing.setdefault("serotype", []).extend(serotype_list)
            if mlst_list:
                existing.setdefault("mlst", []).extend(mlst_list)
        else:
            logger.warning(
                "[%s] Unknown category '%s' for isolate '%s', defaulting to linked",
                current_pmcid, category, isolate_id
            )
            results[current_pmcid]["isolates_with_linking"].append(isolate)

    wb.close()

    # Consolidate accession numbers at PMCID level (deduplicated, order-preserved)  #changed
    for pmcid, data in results.items():  #changed
        merge_acc = data.pop("_merge_accession_number", [])  #changed
        if merge_acc:  #changed
            data["merge_accession_number"] = list(dict.fromkeys(merge_acc))  #changed

    logger.info(
        "Salmonella: processed %d rows across %d PMCIDs",
        row_count, len(results)
    )
    return results


# ===========================================================================
# E.coli converter
# ===========================================================================
def convert_ecoli_isolates(filepath, legend, pmcid):
    """
    Convert a single E.coli isolates Excel file.
    Returns a list of isolate dicts.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]  #changed

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {}
    for i, h in enumerate(headers):
        if h is not None:
            col_idx[h.strip()] = i

    isolates = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        def get_col(name):
            idx = col_idx.get(name)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        isolate_id = safe_string(get_col("Isolate_ID"))
        isolate_source = safe_string(get_col("Isolate Source"))
        isolate_date = extract_years(get_col("Isolate Date"))
        isolate_country = safe_string(get_col("Isolate Country"))
        serotype_raw = safe_string(get_col("Serotype"))
        serotype_list = [serotype_raw] if serotype_raw else []
        mlst_raw = safe_string(get_col("MLST"))
        mlst_list = [mlst_raw] if mlst_raw else []
        ast_raw = get_col("AST_Data")
        spi_list = split_to_list(get_col("SPI"))
        amr_list = split_to_list(get_col("AMR"))
        plasmid_list = split_to_list(get_col("Plasmid"))
        snp_list = split_to_list(get_col("SNP"))
        virulence_list = split_to_list(get_col("Virulence Genes"))

        # E.coli AST is already JSON — parse directly
        serotype_for_ast = serotype_list[0] if serotype_list else None
        ast_data = parse_ast_data(ast_raw, serotype_for_ast, legend, pmcid)

        isolate = {
            "isolate_id": isolate_id,
            "isolate_source": isolate_source,
            "isolate_date": isolate_date,
            "isolate_country": isolate_country,
            "serotype": serotype_list,
            "mlst": mlst_list,
            "ast_data": ast_data,
            "spi": spi_list,
            "amr": amr_list,
            "plasmid": plasmid_list,
            "snp": snp_list,
            "virulence_genes": virulence_list,
        }
        isolates.append(isolate)

    wb.close()
    return isolates


def convert_ecoli_others(filepath):
    """
    Convert a single E.coli others Excel file.
    Returns (isolate_without_linking, no_isolates_only_assayinformation).
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]  #changed

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {}
    for i, h in enumerate(headers):
        if h is not None:
            col_idx[h.strip()] = i

    # Read first data row
    data_row = None
    for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
        data_row = row
        break

    if data_row is None:
        wb.close()
        return [], {}

    def get_col(name):
        idx = col_idx.get(name)
        if idx is None or idx >= len(data_row):
            return None
        return data_row[idx]

    # Parse isolate_without_linking
    iwl_raw = safe_string(get_col("Isolate_without_linking"))
    try:
        iwl = json.loads(iwl_raw) if iwl_raw else []
    except json.JSONDecodeError:
        # Handle Python list syntax with single quotes  #changed
        try:  #changed
            import ast as ast_module  #changed
            iwl = ast_module.literal_eval(iwl_raw) if iwl_raw else []  #changed
        except (ValueError, SyntaxError):  #changed
            iwl = []  #changed
            logger.warning("Failed to parse isolate_without_linking: %s", iwl_raw[:80])  #changed

    # Parse no_isolates_only_assayinformation
    nioa_raw = safe_string(get_col("No_isolates_only_assayinformation"))
    try:
        nioa = json.loads(nioa_raw) if nioa_raw else {}
    except json.JSONDecodeError:
        # Handle Python dict syntax with single quotes  #changed
        try:  #changed
            import ast as ast_module  #changed
            nioa = ast_module.literal_eval(nioa_raw) if nioa_raw else {}  #changed
        except (ValueError, SyntaxError):  #changed
            nioa = {}  #changed
            logger.warning("Failed to parse no_isolates_only_assayinformation: %s", nioa_raw[:80])  #changed

    wb.close()
    return iwl, nioa


def find_ecoli_files(input_dir):
    """
    Discover E.coli file pairs in the input directory.
    Returns a list of (pmcid, isolates_path, others_path) tuples.
    """
    files = os.listdir(input_dir)
    pairs = []
    seen_pmcids = set()

    for f in files:
        if f.endswith(ECOLI_ISOLATES_SUFFIX):
            pmcid = f.replace(ECOLI_ISOLATES_SUFFIX, "")
            if pmcid in seen_pmcids:
                continue
            seen_pmcids.add(pmcid)

            isolates_path = os.path.join(input_dir, f)
            others_path = os.path.join(input_dir, pmcid + ECOLI_OTHERS_SUFFIX)

            if not os.path.exists(others_path):
                logger.warning(
                    "E.coli others file not found for %s (expected: %s)",
                    pmcid, others_path
                )
                others_path = None

            pairs.append((pmcid, isolates_path, others_path))

    return sorted(pairs)


# ===========================================================================
# Validation
# ===========================================================================
def validate_output(pmcid, data):
    """
    Run post-conversion validation checks on a single PMCID output.
    Returns a list of warning strings.
    """
    warnings = []

    for section in ["isolates_with_linking", "isolate_without_linking"]:
        isolates = data.get(section, [])
        if not isinstance(isolates, list):
            continue

        for i, isolate in enumerate(isolates):
            if not isinstance(isolate, dict):  #changed
                continue  #changed
            isolate_label = isolate.get("isolate_id", f"index_{i}")

            for j, ast_entry in enumerate(isolate.get("ast_data", [])):
                for k, ab in enumerate(ast_entry.get("Antibiotics", [])):
                    # Check 1: non-empty name
                    if not ab.get("Name", "").strip():
                        warnings.append(
                            f"[{pmcid}] {isolate_label}: empty antibiotic name at ast_data[{j}].Antibiotics[{k}]"
                        )

                    # Check 2: valid interpretation
                    interp = ab.get("Interpretation", "")
                    if interp not in VALID_INTERP:
                        warnings.append(
                            f"[{pmcid}] {isolate_label}: invalid interpretation '{interp}' for '{ab.get('Name')}'"
                        )

                # Check 3: duplicate antibiotics
                names = [ab.get("Name", "").lower() for ab in ast_entry.get("Antibiotics", [])]
                seen = set()
                for name in names:
                    if name in seen:
                        warnings.append(
                            f"[{pmcid}] {isolate_label}: duplicate antibiotic '{name}'"
                        )
                    seen.add(name)

    return warnings


# ===========================================================================
# Main entry point
# ===========================================================================
def main():
    """Main conversion pipeline."""

    # Create output directory
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    all_warnings = []
    total_pmcids = 0

    # ----- Salmonella -----
    sal_path = os.path.join(INPUT_DIR, SALMONELLA_FILE)
    if os.path.exists(sal_path):
        logger.info("=" * 60)
        logger.info("Processing Salmonella file: %s", sal_path)
        logger.info("=" * 60)

        legend = load_legend(sal_path)
        sal_results = convert_salmonella(sal_path, legend)

        for pmcid, data in sal_results.items():
            # Validate
            warnings = validate_output(pmcid, data)
            all_warnings.extend(warnings)

            # Write JSON
            output_path = os.path.join(OUTPUT_DIR, f"{pmcid}.json")
            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            logger.info("Written: %s (%d isolates)",
                        output_path, len(data["isolates_with_linking"]))
            total_pmcids += 1
    else:
        logger.warning("Salmonella file not found: %s", sal_path)

    # ----- E.coli -----
    ecoli_pairs = find_ecoli_files(INPUT_DIR)
    if ecoli_pairs:
        logger.info("=" * 60)
        logger.info("Processing %d E.coli file pairs", len(ecoli_pairs))
        logger.info("=" * 60)

        # Use Salmonella legend for E.coli as well (abbreviation lookup fallback)
        legend = {}
        sal_path_check = os.path.join(INPUT_DIR, SALMONELLA_FILE)
        if os.path.exists(sal_path_check):
            try:
                legend = load_legend(sal_path_check)
            except ValueError:
                logger.info("No legend available for E.coli abbreviation fallback")

        for pmcid, isolates_path, others_path in ecoli_pairs:
            logger.info("Processing E.coli: %s", pmcid)

            # Convert isolates
            isolates = convert_ecoli_isolates(isolates_path, legend, pmcid)

            # Convert others (if file exists)
            iwl = []
            nioa = {}
            if others_path:
                iwl, nioa = convert_ecoli_others(others_path)

            data = {
                "pmcid": pmcid,
                "isolate_without_linking": iwl,
                "no_isolates_only_assayinformation": nioa,
                "isolates_with_linking": isolates
            }

            # Validate
            warnings = validate_output(pmcid, data)
            all_warnings.extend(warnings)

            # Write JSON
            output_path = os.path.join(OUTPUT_DIR, f"{pmcid}.json")
            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            logger.info("Written: %s (%d isolates)", output_path, len(isolates))
            total_pmcids += 1

    else:
        logger.info("No E.coli file pairs found in %s", INPUT_DIR)

    # ----- Summary -----
    logger.info("=" * 60)
    logger.info("CONVERSION COMPLETE")
    logger.info("  Total PMCIDs processed: %d", total_pmcids)
    logger.info("  Output directory: %s", OUTPUT_DIR)

    if all_warnings:
        logger.info("  Validation warnings: %d", len(all_warnings))
        logger.info("-" * 60)
        for w in all_warnings:
            logger.warning("  VALIDATION: %s", w)
    else:
        logger.info("  Validation warnings: 0 (all clean)")

    logger.info("=" * 60)


if __name__ == "__main__":
    main()
